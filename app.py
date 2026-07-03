from flask import Flask, render_template, request, send_file, jsonify, session, redirect, url_for, flash
from functools import wraps
from contextlib import contextmanager
import duckdb
import io
import os
import re
import json
import csv
import secrets
from hashlib import sha256 as legacy_sha256
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook

app = Flask(__name__)

_secret = os.environ.get("FLASK_SECRET_KEY")
if _secret:
    app.secret_key = _secret
else:
    app.secret_key = secrets.token_hex(32)

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
)

DB_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_DB = "portal_master.db"
DEFAULT_DB_PATH = os.path.join(DB_DIR, MASTER_DB)

DB_NAME_RE = re.compile(r"^[a-zA-Z0-9_\-]+\.db$")


def validate_db_name(name):
    if not name:
        return "Database name is required"
    if not name.endswith(".db"):
        name += ".db"
    if not DB_NAME_RE.match(name):
        return "Only letters, numbers, underscores, and hyphens allowed"
    return None


DISPLAY_COLUMNS = [
    "Salutation",
    "First Name",
    "Last Name",
    "Email",
    "Job Titles",
    "Company Name",
    "Country",
    "Status",
]

COUNTRY_CACHE = None
STATUS_CHOICES = ["New", "Contacted", "Qualified", "Lost"]


def clean_illegal_chars(value):
    if isinstance(value, str):
        return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', value)
    return value


def get_current_db_path():
    db_path = session.get("db_path")
    if not db_path or not os.path.exists(db_path):
        db_path = DEFAULT_DB_PATH
        session["db_path"] = db_path
    return db_path


@contextmanager
def get_db_cxn():
    conn = duckdb.connect(get_current_db_path())
    try:
        ensure_schemas(conn)
        yield conn
    finally:
        conn.close()


def get_users(conn):
    rows = conn.execute("SELECT id, username, role, created_at FROM Users ORDER BY id").fetchall()
    return [dict(zip(["id", "username", "role", "created_at"], [r[0], r[1], r[2], str(r[3])[:10] if r[3] else ""])) for r in rows]


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            if request.is_json or request.path.startswith("/api/"):
                return jsonify({"error": "Authentication required"}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if "user_id" not in session:
                if request.is_json or request.path.startswith("/api/"):
                    return jsonify({"error": "Authentication required"}), 401
                return redirect(url_for("login"))
            if session.get("role") not in roles:
                if request.is_json or request.path.startswith("/api/"):
                    return jsonify({"error": "Insufficient permissions"}), 403
                flash("You do not have permission to perform this action.")
                return redirect(url_for("home"))
            return f(*args, **kwargs)
        return decorated
    return decorator


# ─── AUDIT LOG ───────────────────────────────────────────────────


def log_audit(conn, action, entity_type=None, entity_id=None, details=None):
    max_id = conn.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM AuditLog").fetchone()[0]
    conn.execute("""
        INSERT INTO AuditLog (id, user_id, username, action, entity_type, entity_id, details)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, [
        max_id,
        session.get("user_id"),
        session.get("username"),
        action,
        entity_type,
        str(entity_id) if entity_id is not None else None,
        json.dumps(details) if details else None,
    ])


# ─── LOGIN / LOGOUT ─────────────────────────────────────────────


@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("home"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        with get_db_cxn() as conn:
            user = conn.execute(
                "SELECT id, username, role, password_hash, must_change_pw FROM Users WHERE username = ?",
                [username]
            ).fetchone()

        if user:
            stored_hash = user[3]
            must_change = bool(user[4]) if len(user) > 4 else False
            is_legacy = len(stored_hash) == 64 and all(c in '0123456789abcdef' for c in stored_hash)
            valid = False

            if is_legacy:
                if legacy_sha256(password.encode()).hexdigest() == stored_hash:
                    valid = True
                    new_hash = generate_password_hash(password)
                    with get_db_cxn() as conn:
                        conn.execute("UPDATE Users SET password_hash = ? WHERE id = ?", [new_hash, user[0]])
                        if user[0] == 1:
                            conn.execute("UPDATE Users SET must_change_pw = 1 WHERE id = 1")
                            must_change = True
            else:
                if check_password_hash(stored_hash, password):
                    valid = True

            if valid:
                session["user_id"] = user[0]
                session["username"] = user[1]
                session["role"] = user[2]

                with get_db_cxn() as conn:
                    log_audit(conn, "login", "User", user[0], {"username": user[1]})

                if must_change and user[0] == 1:
                    return redirect(url_for("change_password"))

                return redirect(url_for("home"))

        return render_template("login.html", error="Invalid username or password")
    return render_template("login.html")


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        current = request.form.get("current_password", "").strip()
        new_pw = request.form.get("new_password", "").strip()
        confirm = request.form.get("confirm_password", "").strip()

        if not new_pw or len(new_pw) < 4:
            return render_template("change_password.html", error="Password must be at least 4 characters")
        if new_pw != confirm:
            return render_template("change_password.html", error="Passwords do not match")

        with get_db_cxn() as conn:
            row = conn.execute("SELECT password_hash FROM Users WHERE id = ?", [session["user_id"]]).fetchone()
            if not row:
                return render_template("change_password.html", error="User not found")
            stored_hash = row[0]
            is_legacy = len(stored_hash) == 64 and all(c in '0123456789abcdef' for c in stored_hash)

            if is_legacy:
                if legacy_sha256(current.encode()).hexdigest() != stored_hash:
                    return render_template("change_password.html", error="Current password is incorrect")
            else:
                if not check_password_hash(stored_hash, current):
                    return render_template("change_password.html", error="Current password is incorrect")

            new_hash = generate_password_hash(new_pw)
            conn.execute("UPDATE Users SET password_hash = ?, must_change_pw = 0 WHERE id = ?",
                         [new_hash, session["user_id"]])
            log_audit(conn, "change_password", "User", session["user_id"])

        flash("Password changed successfully.")
        return redirect(url_for("home"))
    return render_template("change_password.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ─── HOME ────────────────────────────────────────────────────────


@app.route("/", methods=["GET"])
@login_required
def home():
    global COUNTRY_CACHE

    tab = request.args.get("tab", "leads")
    export = request.args.get("export")
    selected_countries = request.args.getlist("country")
    status_filter = request.args.get("status", "").strip()
    show_trash = request.args.get("trash", "").strip()

    with get_db_cxn() as conn:
        if COUNTRY_CACHE is None:
            rows = conn.execute("""
                SELECT DISTINCT Country FROM Leads
                WHERE Country IS NOT NULL AND (Status IS NULL OR Status != 'Deleted')
                ORDER BY Country
            """).fetchall()
            COUNTRY_CACHE = [r[0] for r in rows]

        leads_data = None
        total_records = 0
        analytics = None
        trash_data = None
        audit_entries = []

        if tab == "leads" and not show_trash:
            base_query = "FROM Leads WHERE (Status IS NULL OR Status != 'Deleted')"
            params = []
            if selected_countries:
                placeholders = ",".join("?" for _ in selected_countries)
                base_query += f" AND Country IN ({placeholders})"
                params = list(selected_countries)
            if status_filter:
                base_query += " AND Status = ?"
                params.append(status_filter)

            if export:
                export_query = base_query
                export_params = list(params)
                if not selected_countries and not status_filter:
                    export_query = "FROM Leads WHERE (Status IS NULL OR Status != 'Deleted')"
                    export_params = []
                rows = conn.execute(f"SELECT * {export_query}", export_params).fetchall()
                columns = [c[0] for c in conn.description]
                wb = Workbook(write_only=True)
                ws = wb.create_sheet("Filtered Data")
                ws.append(columns)
                for r in rows:
                    ws.append([clean_illegal_chars(v) for v in r])
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                return send_file(
                    output, as_attachment=True, download_name="Filtered_Data.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            query = f"""SELECT {', '.join('"' + c + '"' for c in DISPLAY_COLUMNS)}
                        {base_query} LIMIT 5000"""
            rows = conn.execute(query, params).fetchall()
            leads_data = [dict(zip(DISPLAY_COLUMNS, r)) for r in rows]
            total_records = conn.execute(f"SELECT COUNT(*) {base_query}", params).fetchone()[0]

        elif tab == "leads" and show_trash:
            trash_query = "FROM Leads WHERE Status = 'Deleted'"
            rows = conn.execute(f"SELECT * {trash_query} ORDER BY updated_at DESC LIMIT 5000").fetchall()
            trash_columns = [c[0] for c in conn.description]
            trash_data = [dict(zip(trash_columns, r)) for r in rows]

        elif tab == "analytics":
            analytics = {}
            analytics["countries"] = conn.execute("""
                SELECT Country, COUNT(*) AS cnt FROM Leads
                WHERE Country IS NOT NULL AND (Status IS NULL OR Status != 'Deleted')
                GROUP BY Country ORDER BY cnt DESC
            """).fetchall()
            analytics["job_titles"] = conn.execute("""
                SELECT "Job Titles", COUNT(*) AS cnt FROM Leads
                WHERE "Job Titles" IS NOT NULL AND "Job Titles" != ''
                AND (Status IS NULL OR Status != 'Deleted')
                GROUP BY "Job Titles" ORDER BY cnt DESC LIMIT 15
            """).fetchall()
            analytics["status_breakdown"] = conn.execute("""
                SELECT COALESCE(Status, 'New') AS st, COUNT(*) AS cnt FROM Leads
                GROUP BY st ORDER BY cnt DESC
            """).fetchall()
            analytics["over_time"] = conn.execute("""
                SELECT DATE(created_at) AS day, COUNT(*) AS cnt FROM Leads
                WHERE created_at IS NOT NULL AND (Status IS NULL OR Status != 'Deleted')
                GROUP BY day ORDER BY day
            """).fetchall()

        elif tab == "activity" and session.get("role") == "admin":
            raw = conn.execute("""
                SELECT id, user_id, username, action, entity_type, entity_id, details, created_at
                FROM AuditLog ORDER BY created_at DESC LIMIT 200
            """).fetchall()
            audit_entries = []
            for r in raw:
                item = dict(zip(["id", "user_id", "username", "action", "entity_type", "entity_id", "details", "created_at"], r))
                if item["created_at"]:
                    item["created_at"] = str(item["created_at"])[:19]
                else:
                    item["created_at"] = ""
                audit_entries.append(item)

        users = get_users(conn)

    db_name = os.path.basename(get_current_db_path())
    db_list = [f for f in os.listdir(DB_DIR) if f.endswith(".db") and os.path.isfile(os.path.join(DB_DIR, f))]
    db_list.sort()

    return render_template(
        "home.html",
        tab=tab,
        campaigns=leads_data,
        countries=COUNTRY_CACHE,
        selected_countries=selected_countries,
        total_records=total_records,
        users=users,
        analytics=analytics,
        username=session.get("username"),
        user_role=session.get("role"),
        current_db=db_name,
        db_path=get_current_db_path(),
        db_list=db_list,
        status_choices=STATUS_CHOICES,
        selected_status=status_filter,
        show_trash=show_trash,
        trash_data=trash_data,
        audit_entries=audit_entries,
    )


# ─── ADD LEAD (form) ────────────────────────────────────────────


@app.route("/add-lead", methods=["POST"])
@login_required
@role_required("admin")
def add_lead():
    salutation = request.form.get("salutation", "").strip()
    first_name = request.form.get("first_name", "").strip()
    last_name = request.form.get("last_name", "").strip()
    email = request.form.get("email", "").strip()
    job_titles = request.form.get("job_titles", "").strip()
    company_name = request.form.get("company_name", "").strip()
    country = request.form.get("country", "").strip()

    if not first_name or not last_name:
        return jsonify({"error": "First Name and Last Name are required"}), 400

    with get_db_cxn() as conn:
        conn.execute("""
            INSERT INTO Leads ("Salutation", "First Name", "Last Name", "Email", "Job Titles", "Company Name", "Country", "Status")
            VALUES (?, ?, ?, ?, ?, ?, ?, 'New')
        """, [salutation, first_name, last_name, email, job_titles, company_name, country])
        log_audit(conn, "create", "Lead", None, {"first_name": first_name, "last_name": last_name, "email": email})
    return jsonify({"success": True})


# ─── USER MANAGEMENT ────────────────────────────────────────────


@app.route("/users/create", methods=["POST"])
@login_required
@role_required("admin")
def create_user():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    role = request.form.get("role", "viewer").strip()

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400
    if len(password) < 4:
        return jsonify({"error": "Password must be at least 4 characters"}), 400

    pw_hash = generate_password_hash(password)
    try:
        with get_db_cxn() as conn:
            max_id = conn.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM Users").fetchone()[0]
            conn.execute("INSERT INTO Users (id, username, password_hash, role) VALUES (?, ?, ?, ?)",
                         [max_id, username, pw_hash, role])
            log_audit(conn, "create", "User", max_id, {"username": username, "role": role})
        return jsonify({"success": True})
    except Exception as e:
        error_msg = str(e)
        if "UNIQUE" in error_msg or "unique" in error_msg.lower():
            return jsonify({"error": f"Username '{username}' already exists"}), 409
        return jsonify({"error": error_msg}), 400


@app.route("/users/delete/<int:user_id>", methods=["POST"])
@login_required
@role_required("admin")
def delete_user(user_id):
    with get_db_cxn() as conn:
        user = conn.execute("SELECT id, username FROM Users WHERE id = ?", [user_id]).fetchone()
        if not user:
            return jsonify({"error": "User not found"}), 404
        if user_id == 1:
            return jsonify({"error": "Cannot delete the default admin"}), 403
        conn.execute("DELETE FROM Users WHERE id = ?", [user_id])
        log_audit(conn, "delete", "User", user_id, {"username": user[1]})
    return jsonify({"success": True})


# ─── DATABASE MANAGEMENT ──────────────────────────────────────────

LEADS_SCHEMA = """
CREATE TABLE IF NOT EXISTS Leads (
    "Salutation" VARCHAR,
    "First Name" VARCHAR,
    "Last Name" VARCHAR,
    "Email" VARCHAR,
    "Job Titles" VARCHAR,
    "Company Name" VARCHAR,
    "Country" VARCHAR
)
"""

USERS_SCHEMA = """
CREATE TABLE IF NOT EXISTS Users (
    id INTEGER PRIMARY KEY,
    username VARCHAR UNIQUE,
    password_hash VARCHAR,
    role VARCHAR DEFAULT 'viewer',
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)
"""

AUDIT_LOG_SCHEMA = """
CREATE TABLE IF NOT EXISTS AuditLog (
    id INTEGER PRIMARY KEY,
    user_id INTEGER,
    username VARCHAR,
    action VARCHAR,
    entity_type VARCHAR,
    entity_id VARCHAR,
    details VARCHAR,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)
"""


def ensure_schemas(conn):
    conn.execute(LEADS_SCHEMA)
    conn.execute(USERS_SCHEMA)
    conn.execute(AUDIT_LOG_SCHEMA)

    migrations = [
        ('Leads', 'Status', "VARCHAR DEFAULT 'New'"),
        ('Leads', 'created_at', "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
        ('Leads', 'updated_at', "TIMESTAMP"),
        ('Users', 'must_change_pw', "INTEGER DEFAULT 0"),
    ]
    for table, col, typ in migrations:
        try:
            conn.execute(f'ALTER TABLE "{table}" ADD COLUMN IF NOT EXISTS "{col}" {typ}')
        except Exception:
            pass

    existing = conn.execute("SELECT COUNT(*) FROM Users").fetchone()[0]
    if existing == 0:
        admin_hash = generate_password_hash("admin")
        conn.execute("INSERT INTO Users (id, username, password_hash, role, must_change_pw) VALUES (1, 'admin', ?, 'admin', 1)",
                     [admin_hash])
        viewer_hash = generate_password_hash("viewer")
        conn.execute("INSERT INTO Users (id, username, password_hash, role) VALUES (2, 'viewer', ?, 'viewer')",
                     [viewer_hash])


@app.route("/api/databases", methods=["GET"])
@login_required
def api_list_databases():
    files = [f for f in os.listdir(DB_DIR) if f.endswith(".db") and os.path.isfile(os.path.join(DB_DIR, f))]
    files.sort(key=lambda f: os.path.getmtime(os.path.join(DB_DIR, f)), reverse=True)
    current = os.path.basename(get_current_db_path())
    return jsonify({
        "databases": files,
        "current": current,
        "path": get_current_db_path(),
    })


@app.route("/api/databases/create", methods=["POST"])
@login_required
@role_required("admin")
def api_create_database():
    global COUNTRY_CACHE
    body = request.get_json(silent=True)
    name = ((body or {}).get("name") or "").strip()
    err = validate_db_name(name)
    if err:
        return jsonify({"error": err}), 400
    if not name.endswith(".db"):
        name += ".db"
    path = os.path.join(DB_DIR, name)
    if os.path.exists(path):
        return jsonify({"error": f"Database '{name}' already exists"}), 409
    try:
        conn = duckdb.connect(path)
        ensure_schemas(conn)
        conn.close()
        with get_db_cxn() as conn2:
            log_audit(conn2, "create", "Database", name)
        return jsonify({"success": True, "name": name, "path": path})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/databases/switch", methods=["POST"])
@login_required
@role_required("admin")
def api_switch_database():
    global COUNTRY_CACHE
    body = request.get_json(silent=True)
    name = ((body or {}).get("name") or "").strip()
    if not name:
        return jsonify({"error": "Database name is required"}), 400
    path = os.path.join(DB_DIR, name if name.endswith(".db") else name + ".db")
    if not os.path.exists(path):
        return jsonify({"error": f"Database '{name}' not found"}), 404
    old_path = session.get("db_path", DEFAULT_DB_PATH)
    session["db_path"] = path
    COUNTRY_CACHE = None
    with get_db_cxn() as conn:
        log_audit(conn, "switch", "Database", name, {"from": old_path})
    return jsonify({"success": True, "name": name, "path": path})


@app.route("/api/databases/delete", methods=["POST"])
@login_required
@role_required("admin")
def api_delete_database():
    global COUNTRY_CACHE
    body = request.get_json(silent=True)
    name = ((body or {}).get("name") or "").strip()
    if not name:
        return jsonify({"error": "Database name is required"}), 400
    path = os.path.join(DB_DIR, name if name.endswith(".db") else name + ".db")
    if not os.path.exists(path):
        return jsonify({"error": f"Database '{name}' not found"}), 404
    if path == get_current_db_path():
        return jsonify({"error": "Cannot delete the active database"}), 400
    if os.path.basename(path) == MASTER_DB:
        return jsonify({"error": "Cannot delete the master database"}), 403
    os.remove(path)
    with get_db_cxn() as conn:
        log_audit(conn, "delete", "Database", name)
    return jsonify({"success": True})


@app.route("/api/databases/rename", methods=["POST"])
@login_required
@role_required("admin")
def api_rename_database():
    global COUNTRY_CACHE
    body = request.get_json(silent=True)
    old_name = ((body or {}).get("old_name") or "").strip()
    new_name = ((body or {}).get("new_name") or "").strip()
    if not old_name or not new_name:
        return jsonify({"error": "Both old_name and new_name are required"}), 400
    if old_name == MASTER_DB:
        return jsonify({"error": "Cannot rename the master database"}), 403
    err = validate_db_name(new_name)
    if err:
        return jsonify({"error": err}), 400
    if not new_name.endswith(".db"):
        new_name += ".db"
    old_path = os.path.join(DB_DIR, old_name if old_name.endswith(".db") else old_name + ".db")
    new_path = os.path.join(DB_DIR, new_name)
    if not os.path.exists(old_path):
        return jsonify({"error": f"Database '{old_name}' not found"}), 404
    if os.path.exists(new_path):
        return jsonify({"error": f"Database '{new_name}' already exists"}), 409
    try:
        os.rename(old_path, new_path)
        current = get_current_db_path()
        if current == old_path:
            session["db_path"] = new_path
            COUNTRY_CACHE = None
        with get_db_cxn() as conn:
            log_audit(conn, "rename", "Database", new_name, {"from": old_name})
        return jsonify({"success": True, "name": new_name, "path": new_path})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── API ENDPOINTS ───────────────────────────────────────────────

API_COLUMNS = [
    ("salutation", "Salutation"),
    ("first_name", "First Name"),
    ("last_name", "Last Name"),
    ("email", "Email"),
    ("job_titles", "Job Titles"),
    ("company_name", "Company Name"),
    ("country", "Country"),
    ("status", "Status"),
]


@app.route("/api/leads", methods=["GET"])
@login_required
def api_get_leads():
    limit = request.args.get("limit", 50, type=int)
    offset = request.args.get("offset", 0, type=int)
    country_q = request.args.get("country", "").strip()
    search = request.args.get("search", "").strip()
    sort = request.args.get("sort", "first_name").strip()
    dir_ = request.args.get("dir", "asc").strip()
    status_q = request.args.get("status", "").strip()
    trash = request.args.get("trash", "").strip()

    if sort not in {k for k, v in API_COLUMNS}:
        sort = "first_name"
    dir_ = "DESC" if dir_.lower() == "desc" else "ASC"

    db_cols = ", ".join(f'"{v}"' for _, v in API_COLUMNS)

    with get_db_cxn() as conn:
        if trash:
            query = "FROM Leads WHERE Status = 'Deleted'"
        else:
            query = "FROM Leads WHERE (Status IS NULL OR Status != 'Deleted')"

        if country_q:
            query += " AND Country = ?"
        if search:
            query += ' AND ("First Name" ILIKE ? OR "Last Name" ILIKE ? OR Email ILIKE ? OR "Company Name" ILIKE ?)'
            pat = f"%{search}%"
        if status_q and not trash:
            query += " AND Status = ?"

        count_q = f"SELECT COUNT(*) {query}"
        params_list = []
        p = []
        if country_q:
            p.append(country_q)
        if search:
            p.extend([pat, pat, pat, pat])
        if status_q and not trash:
            p.append(status_q)
        row = conn.execute(count_q, p).fetchone()
        total = row[0]

        sort_db = dict(API_COLUMNS).get(sort, "First Name")
        sql = f"SELECT rowid, {db_cols} {query} ORDER BY \"{sort_db}\" {dir_} LIMIT ? OFFSET ?"
        rows = conn.execute(sql, p + [limit, offset]).fetchall()

        keys = ["id"] + [k for k, v in API_COLUMNS]
        data = [dict(zip(keys, r)) for r in rows]

    return jsonify({"total": total, "offset": offset, "limit": limit, "data": data})


@app.route("/api/leads", methods=["POST"])
@login_required
@role_required("admin")
def api_create_lead():
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "Request body must be JSON"}), 400
    salutation = (body.get("salutation") or "").strip()
    first_name = (body.get("first_name") or body.get("firstName") or "").strip()
    last_name = (body.get("last_name") or body.get("lastName") or "").strip()
    email = (body.get("email") or "").strip()
    job_titles = (body.get("job_titles") or body.get("jobTitles") or "").strip()
    company_name = (body.get("company_name") or body.get("companyName") or "").strip()
    country = (body.get("country") or "").strip()
    if not first_name or not last_name:
        return jsonify({"error": "first_name and last_name are required"}), 400
    with get_db_cxn() as conn:
        conn.execute("""
            INSERT INTO Leads ("Salutation", "First Name", "Last Name", "Email", "Job Titles", "Company Name", "Country", "Status")
            VALUES (?, ?, ?, ?, ?, ?, ?, 'New')
        """, [salutation, first_name, last_name, email, job_titles, company_name, country])
        log_audit(conn, "create", "Lead", None, {"first_name": first_name, "last_name": last_name, "email": email})
    return jsonify({"success": True, "message": "Lead created"}), 201


@app.route("/api/leads/countries", methods=["GET"])
@login_required
def api_countries():
    with get_db_cxn() as conn:
        rows = conn.execute("""
            SELECT Country, COUNT(*) AS cnt FROM Leads
            WHERE Country IS NOT NULL AND (Status IS NULL OR Status != 'Deleted')
            GROUP BY Country ORDER BY cnt DESC
        """).fetchall()
    return jsonify([{"country": r[0], "count": r[1]} for r in rows])


@app.route("/api/leads/stats", methods=["GET"])
@login_required
def api_stats():
    with get_db_cxn() as conn:
        total = conn.execute("SELECT COUNT(*) FROM Leads WHERE (Status IS NULL OR Status != 'Deleted')").fetchone()[0]
        with_email = conn.execute("SELECT COUNT(*) FROM Leads WHERE Email IS NOT NULL AND Email != '' AND (Status IS NULL OR Status != 'Deleted')").fetchone()[0]
        countries = conn.execute("""
            SELECT Country, COUNT(*) AS cnt FROM Leads
            WHERE Country IS NOT NULL AND (Status IS NULL OR Status != 'Deleted')
            GROUP BY Country ORDER BY cnt DESC
        """).fetchall()
        salutations = conn.execute("""
            SELECT Salutation, COUNT(*) AS cnt FROM Leads
            WHERE Salutation IS NOT NULL AND Salutation != ''
            AND (Status IS NULL OR Status != 'Deleted')
            GROUP BY Salutation ORDER BY cnt DESC
        """).fetchall()
    return jsonify({
        "total": total,
        "with_email": with_email,
        "email_fill_rate": round(with_email / total * 100, 1) if total else 0,
        "countries": [{"country": r[0], "count": r[1]} for r in countries],
        "salutations": [{"salutation": r[0], "count": r[1]} for r in salutations],
    })


# ─── LEAD UPDATE / DELETE / RESTORE ─────────────────────────────

LEAD_FIELD_MAP = {
    "salutation": "Salutation",
    "first_name": "First Name",
    "last_name": "Last Name",
    "email": "Email",
    "job_titles": "Job Titles",
    "company_name": "Company Name",
    "country": "Country",
    "status": "Status",
}


@app.route("/api/leads/<int:rowid>", methods=["PUT"])
@login_required
@role_required("admin")
def api_update_lead(rowid):
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "Request body must be JSON"}), 400

    set_clauses = []
    params = []
    for json_key, db_col in LEAD_FIELD_MAP.items():
        if json_key in body:
            set_clauses.append(f'"{db_col}" = ?')
            params.append((body.get(json_key) or "").strip())

    if not set_clauses:
        return jsonify({"error": "No fields to update"}), 400

    set_clauses.append('"updated_at" = CURRENT_TIMESTAMP')
    params.append(rowid)
    sql = f'UPDATE Leads SET {", ".join(set_clauses)} WHERE rowid = ?'
    with get_db_cxn() as conn:
        conn.execute(sql, params)
        log_audit(conn, "update", "Lead", rowid, body)
    return jsonify({"success": True})


@app.route("/api/leads/<int:rowid>", methods=["DELETE"])
@login_required
@role_required("admin")
def api_delete_lead(rowid):
    with get_db_cxn() as conn:
        conn.execute('UPDATE Leads SET "Status" = \'Deleted\', "updated_at" = CURRENT_TIMESTAMP WHERE rowid = ?', [rowid])
        log_audit(conn, "delete", "Lead", rowid)
    return jsonify({"success": True})


@app.route("/api/leads/<int:rowid>/restore", methods=["POST"])
@login_required
@role_required("admin")
def api_restore_lead(rowid):
    with get_db_cxn() as conn:
        conn.execute('UPDATE Leads SET "Status" = \'New\', "updated_at" = CURRENT_TIMESTAMP WHERE rowid = ?', [rowid])
        log_audit(conn, "restore", "Lead", rowid)
    return jsonify({"success": True})


@app.route("/api/leads/bulk-delete", methods=["POST"])
@login_required
@role_required("admin")
def api_bulk_delete():
    body = request.get_json(silent=True)
    if not body or not isinstance(body.get("ids"), list):
        return jsonify({"error": "Provide an ids array"}), 400
    ids = body["ids"]
    if not ids:
        return jsonify({"error": "No ids provided"}), 400
    with get_db_cxn() as conn:
        placeholders = ",".join("?" for _ in ids)
        conn.execute(f'UPDATE Leads SET "Status" = \'Deleted\', "updated_at" = CURRENT_TIMESTAMP WHERE rowid IN ({placeholders})', ids)
        log_audit(conn, "bulk_delete", "Lead", None, {"ids": ids, "count": len(ids)})
    return jsonify({"success": True, "deleted": len(ids)})


# ─── BULK UPLOAD ──────────────────────────────────────────────────

ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}


@app.route("/bulk-upload", methods=["POST"])
@login_required
@role_required("admin")
def bulk_upload():
    inserted = 0
    errors = []

    body = request.get_json(silent=True)
    if body and body.get("filetype") == "csv":
        try:
            with get_db_cxn() as conn:
                content = body.get("csv", "")
                reader = csv.DictReader(io.StringIO(content))
                header = [h.strip().lower().replace(" ", "_") for h in reader.fieldnames] if reader.fieldnames else []
                col_map = _map_columns(header)
                if not col_map.get("first_name") or not col_map.get("last_name"):
                    return jsonify({"error": "CSV must contain First Name and Last Name columns"}), 400

                for i, row_dict in enumerate(reader, 2):
                    if not any(row_dict.values()):
                        continue
                    _insert_lead(conn, col_map, row_dict, i, errors, inserted)
                    if not errors or errors[-1].get("row") != i:
                        inserted += 1
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        return jsonify({"success": True, "inserted": inserted, "errors": errors[:10], "total_errors": len(errors)})

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files["file"]
    if f.filename == "":
        return jsonify({"error": "Empty filename"}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": "Only .xlsx, .xls, .csv files are allowed"}), 400

    try:
        with get_db_cxn() as conn:
            if ext == ".csv":
                content = f.read().decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(content))
                header = [h.strip().lower().replace(" ", "_") for h in reader.fieldnames] if reader.fieldnames else []
                col_map = _map_columns(header)
                if not col_map.get("first_name") or not col_map.get("last_name"):
                    return jsonify({"error": "CSV must contain First Name and Last Name columns"}), 400

                for i, row_dict in enumerate(reader, 2):
                    if not any(row_dict.values()):
                        continue
                    _insert_lead(conn, col_map, row_dict, i, errors, inserted)
                    if not errors or errors[-1].get("row") != i:
                        inserted += 1
            else:
                from openpyxl import load_workbook
                wb = load_workbook(f, read_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                header_row = next(rows_iter, [])
                header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in header_row]
                col_map = _map_columns(header)
                if not col_map.get("first_name") or not col_map.get("last_name"):
                    return jsonify({"error": "Excel must contain First Name and Last Name columns"}), 400
                for i, row in enumerate(rows_iter, 2):
                    vals = [str(v).strip() if v is not None else "" for v in row]
                    row_dict = dict(zip(header, vals))
                    _insert_lead(conn, col_map, row_dict, i, errors, inserted)
                    if not errors or errors[-1].get("row") != i:
                        inserted += 1

            if inserted > 0:
                log_audit(conn, "bulk_upload", "Lead", None, {"inserted": inserted, "errors": len(errors)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify({
        "success": True,
        "inserted": inserted,
        "errors": errors[:10],
        "total_errors": len(errors),
    })


COLUMN_ALIASES = {
    "salutation": "Salutation",
    "first_name": "First Name",
    "last_name": "Last Name",
    "email": "Email",
    "job_titles": "Job Titles",
    "job_title": "Job Titles",
    "company_name": "Company Name",
    "company": "Company Name",
    "country": "Country",
}


def _map_columns(header):
    mapping = {}
    for h in header:
        h_clean = h.strip().replace(" ", "_").lower()
        if h_clean in COLUMN_ALIASES:
            mapping[h_clean] = COLUMN_ALIASES[h_clean]
    return mapping


def _insert_lead(conn, col_map, row_dict, row_num, errors, inserted):
    try:
        rev = {v: k for k, v in col_map.items()}
        sal = row_dict.get(rev.get("Salutation", ""), "")
        fn = row_dict.get(rev.get("First Name", ""), "")
        ln = row_dict.get(rev.get("Last Name", ""), "")
        em = row_dict.get(rev.get("Email", ""), "")
        jt = row_dict.get(rev.get("Job Titles", ""), "")
        co = row_dict.get(rev.get("Company Name", ""), "")
        cn = row_dict.get(rev.get("Country", ""), "")
        if not fn or not ln:
            errors.append({"row": row_num, "error": "Missing First Name or Last Name"})
            return

        if em:
            existing = conn.execute(
                "SELECT COUNT(*) FROM Leads WHERE LOWER(Email) = LOWER(?) AND (Status IS NULL OR Status != 'Deleted')",
                [em]
            ).fetchone()[0]
            if existing > 0:
                errors.append({"row": row_num, "error": "Duplicate email, skipped"})
                return

        conn.execute("""
            INSERT INTO Leads ("Salutation", "First Name", "Last Name", "Email", "Job Titles", "Company Name", "Country", "Status")
            VALUES (?, ?, ?, ?, ?, ?, ?, 'New')
        """, [sal, fn, ln, em, jt, co, cn])
    except Exception as e:
        errors.append({"row": row_num, "error": str(e)})


# ─── AUDIT LOG API ──────────────────────────────────────────────


@app.route("/api/audit-log", methods=["GET"])
@login_required
@role_required("admin")
def api_audit_log():
    limit = request.args.get("limit", 200, type=int)
    with get_db_cxn() as conn:
        rows = conn.execute("""
            SELECT id, user_id, username, action, entity_type, entity_id, details, created_at
            FROM AuditLog ORDER BY created_at DESC LIMIT ?
        """, [limit]).fetchall()
        keys = ["id", "user_id", "username", "action", "entity_type", "entity_id", "details", "created_at"]
        data = []
        for r in rows:
            d = dict(zip(keys, r))
            d["created_at"] = str(d["created_at"])[:19] if d["created_at"] else ""
            data.append(d)
    return jsonify({"data": data})


if __name__ == "__main__":
    app.run(host='127.0.0.1', port=5001, debug=False, threaded=True)
